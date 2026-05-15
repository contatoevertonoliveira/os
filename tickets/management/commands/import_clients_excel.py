from django.core.management.base import BaseCommand
import csv
import os

from tickets.client_import import ClientImporter

from openpyxl import load_workbook

class Command(BaseCommand):
    help = 'Importa/atualiza clientes a partir de data/clientes.csv (preferencial) ou data/media clientes.xlsx'

    def handle(self, *args, **options):
        dry_run = bool(options.get('dry_run'))

        file_path = 'data/clientes.csv'
        file_kind = 'csv'
        if not os.path.exists(file_path):
            file_path = 'data/clientes.xlsx'
            file_kind = 'xlsx'
        if not os.path.exists(file_path):
            file_path = 'media/clientes.xlsx'
            file_kind = 'xlsx'

        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR('Arquivo não encontrado: data/clientes.csv ou data/media clientes.xlsx'))
            return

        try:
            self.stdout.write(self.style.SUCCESS(f'Reading from {file_path}...'))
            if file_kind == 'csv':
                rows = self.read_csv_best_effort(file_path)
            else:
                rows = self.read_xlsx(file_path)
            
            self.stdout.write(self.style.SUCCESS(f'Found {len(rows)} rows to process.'))

            importer = ClientImporter(stdout=self.stdout)
            result = importer.import_rows(rows, dry_run=dry_run)
            self.stdout.write(self.style.SUCCESS(f"Processados: {result['processed']} | Criados: {result['created']} | Atualizados: {result['updated']}"))

            self.stdout.write(self.style.SUCCESS('Import completed successfully.'))

        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Critical error: {e}'))

    def add_arguments(self, parser):
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Lê e valida o arquivo, mas não grava no banco.',
        )

    def read_csv_best_effort(self, file_path):
        for encoding in ('utf-8-sig', 'utf-8', 'latin-1'):
            try:
                with open(file_path, 'r', encoding=encoding, newline='') as f:
                    reader = csv.DictReader(f, delimiter=';')
                    return [row for row in reader]
            except UnicodeDecodeError:
                continue
        raise RuntimeError('Não foi possível decodificar o arquivo CSV.')

    def read_xlsx(self, file_path):
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            return []
        headers = [str(h).strip() if h is not None else '' for h in headers]
        result = []
        for values in rows_iter:
            if values is None:
                continue
            row = {}
            for i, header in enumerate(headers):
                if not header:
                    continue
                row[header] = values[i] if i < len(values) else None
            result.append(row)
        return result
