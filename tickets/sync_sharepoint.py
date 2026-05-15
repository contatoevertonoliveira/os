import threading
from datetime import timedelta
from io import BytesIO

from django.db.models import Q
from django.utils import timezone

from openpyxl import load_workbook

from .client_import import ClientImporter
from .models import ClientSyncState
from .microsoft_graph import (
    MicrosoftAuthError,
    download_sharepoint_driveitem_content,
    fetch_sharepoint_driveitem_metadata,
    get_graph_access_token,
    shared_clients_url,
)


def parse_xlsx_bytes(xlsx_bytes):
    wb = load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return []
    headers = [str(h).strip() if h is not None else '' for h in headers]
    result = []
    for values in rows_iter:
        if values is None:
            continue
        row = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            row[header] = values[i] if i < len(values) else None
        result.append(row)
    return result


def get_state():
    state, _ = ClientSyncState.objects.get_or_create(source='sharepoint')
    return state


def acquire_lock(state):
    stale_before = timezone.now() - timedelta(minutes=15)
    updated = (
        ClientSyncState.objects.filter(pk=state.pk)
        .filter(Q(is_running=False) | Q(is_running=True, running_started_at__lt=stale_before))
        .update(is_running=True, running_started_at=timezone.now())
    )
    return updated > 0


def release_lock(state):
    ClientSyncState.objects.filter(pk=state.pk).update(is_running=False)


def sync_clients_from_sharepoint(stdout=None, dry_run=False, force=False):
    state = get_state()
    if not acquire_lock(state):
        return {'status': 'skipped', 'reason': 'running'}

    try:
        url = shared_clients_url()
        if not url:
            state.last_error = 'CLIENTS_SHAREPOINT_URL não configurado.'
            state.last_checked_at = timezone.now()
            state.save()
            return {'status': 'error', 'error': state.last_error}

        access_token = get_graph_access_token()
        if not access_token:
            state.last_error = 'Autenticação Microsoft necessária.'
            state.last_checked_at = timezone.now()
            state.save()
            return {'status': 'requires_auth'}

        meta = fetch_sharepoint_driveitem_metadata(url, access_token)
        state.last_checked_at = timezone.now()
        state.remote_last_modified = None
        if meta and meta.get('lastModifiedDateTime'):
            try:
                from datetime import datetime
                state.remote_last_modified = datetime.fromisoformat(meta['lastModifiedDateTime'].replace('Z', '+00:00'))
            except Exception:
                state.remote_last_modified = None

        remote_etag = meta.get('eTag') if meta else None
        if not force and remote_etag and state.etag and remote_etag == state.etag:
            state.save()
            return {'status': 'no_change', 'etag': remote_etag}

        xlsx_bytes = download_sharepoint_driveitem_content(url, access_token)
        rows = parse_xlsx_bytes(xlsx_bytes)
        importer = ClientImporter(stdout=stdout)
        result = importer.import_rows(rows, dry_run=dry_run)

        if not dry_run:
            state.etag = remote_etag or state.etag
            state.last_synced_at = timezone.now()
            state.last_success_at = timezone.now()
            state.last_error = None
        state.save()
        return {'status': 'ok', 'result': result, 'etag': remote_etag}
    except MicrosoftAuthError as e:
        state.last_error = str(e)
        state.save()
        return {'status': 'requires_auth', 'error': str(e)}
    except Exception as e:
        state.last_error = str(e)
        state.save()
        return {'status': 'error', 'error': str(e)}
    finally:
        release_lock(state)


def trigger_sync_background():
    t = threading.Thread(target=sync_clients_from_sharepoint, kwargs={'dry_run': False, 'force': False}, daemon=True)
    t.start()


def maybe_trigger_sync(request):
    if not request.user.is_authenticated:
        return
    url = shared_clients_url()
    if not url:
        return

    key = 'clients_sharepoint_sync_checked'
    if request.session.get(key):
        return

    state = get_state()
    now = timezone.now()
    threshold = now - timedelta(minutes=10)
    last = state.last_checked_at or state.last_synced_at
    if last and last >= threshold:
        request.session[key] = True
        return

    request.session[key] = True
    trigger_sync_background()
